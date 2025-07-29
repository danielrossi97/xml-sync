import pandas as pd
import os
from openpyxl import Workbook

def carregar_registros_existentes(caminho_excel):
    if os.path.exists(caminho_excel):
        return pd.read_excel(caminho_excel)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["CNPJ", "NomeArquivo", "DataModificacao", "DataTransferencia"])
        wb.save(caminho_excel)
        return pd.DataFrame(columns=["CNPJ", "NomeArquivo", "DataModificacao", "DataTransferencia"])

def registrar_transferencias(caminho_excel, registros):
    df_existente = carregar_registros_existentes(caminho_excel)
    df_novo = pd.DataFrame(registros, columns=["CNPJ", "NomeArquivo", "DataModificacao", "DataTransferencia"])
    df_novo["CNPJ"] = df_novo["CNPJ"].astype(str)
    df_final = pd.concat([df_existente, df_novo], ignore_index=True)
    
    # Salva o Excel
    df_final.to_excel(caminho_excel, index=False)

    # Ajusta largura das colunas
    from openpyxl import load_workbook

    wb = load_workbook(caminho_excel)
    ws = wb.active

    largura_colunas = {
        "A": 18,   # CNPJ
        "B": 70,   # NomeArquivo
        "C": 22,   # DataModificacao
        "D": 22    # DataTransferencia
    }

    for col, largura in largura_colunas.items():
        ws.column_dimensions[col].width = largura
    
    wb.save(caminho_excel)